from openpyxl import Workbook
import sys


def write_results_excel_file(name, alg, pars, results):
    """Write the simulation results in an output excel file"""
    # Create excel file with the corresponding sheets
    sheets = []
    wb = Workbook()
    sheets.append(wb.active)
    sheets[0].title = 'SIMULATION RESULTS'
    # Define column width
    for column in range(1, 11):
        column_char = str(chr(64 + column))
        if column == 2:
            sheets[0].column_dimensions[column_char].width = 60
        else:
            sheets[0].column_dimensions[column_char].width = 20
    # Write headers
    header = ['Algorithm', 'Params', 'Unscaled Model Time', 'Unscaled Predict Time', 'Unscaled Train Score',
              'Unscaled Test Score', 'Scaled Model Time', 'Scaled Predict Time', 'Scaled Train Score',
              'Scaled Test Score']
    for i in range(len(header)):
        sheets[0].cell(1, i + 1).value = header[i]
    # Write algorithms
    for i in range(len(alg)):
        sheets[0].cell(i + 2, 1).value = alg[i]
    # Write parameters
    for i in range(len(pars)):
        str_params = ''
        for key, value in pars[i].items():
            str_params += ' ' + key + '=' + str(value)
        sheets[0].cell(i + 2, 2).value = str_params
    # Write data in excel sheet
    for i in range(results.shape[0]):
        for j in range(results.shape[1]):
            sheets[0].cell(i + 2, j + 3).value = results[i, j]
    try:
        wb.save(name)
    except PermissionError:
        sys.exit('ERROR: Excel file open. Please close it to be modified')
