import numpy as np
import time
import sys
from openpyxl import Workbook
import matplotlib.pyplot as plt
from sklearn.neighbors import KNeighborsClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.svm import LinearSVC
from sklearn.naive_bayes import GaussianNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.svm import SVC
from sklearn.neural_network import MLPClassifier
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import MinMaxScaler


class SupervisedAlgorithms:
    """Class to operate with a dataset in CSV format"""

    def __init__(self):
        self.fig_width = 20
        self.fig_height = 10
        self.bar_width = 0.25
        self.method = []
        self.parameters = []
        self.out_data = np.zeros([0, 8])
        self.pca = None
        self.X_train = None
        self.X_train_scaled = None
        self.X_train_scaled_pca = None
        self.X_test_scaled_pca = None
        self.X_train_scaled_pca_output0 = None
        self.X_train_scaled_pca_output1 = None
        self.X_test = None
        self.X_test_scaled = None
        self.y_train = None
        self.y_test = None
        self.y_train_output0 = None
        self.y_train_output1 = None
        self.y_test_output0 = None
        self.y_test_output1 = None

    def train_test_split(self, feature_data, class_data, test_size, algorithm):
        """Split data into training and test datasets and plot the class distribution in each set"""
        self.X_train, self.X_test, self.y_train, self.y_test = train_test_split(feature_data, class_data,
                                                                                test_size=test_size, shuffle=True,
                                                                                stratify=class_data, random_state=1)
        self.y_train = np.ravel(self.y_train)
        self.y_test = np.ravel(self.y_test)
        print("X_train type: {} and shape: {}".format(type(self.X_train), self.X_train.shape))
        print("X_test type: {} and shape: {}".format(type(self.X_test), self.X_test.shape))
        print("y_train type: {} and shape: {}".format(type(self.y_train), self.y_train.shape))
        print("y_test type: {} and shape: {} \n".format(type(self.y_test), self.y_test.shape))
        self.data_scaling(algorithm)
        self.plot_output_class_distribution()
        return self.X_train_scaled, self.y_train

    def plot_output_class_distribution(self):
        """Plot the class distribution in the training and test dataset"""
        self.y_train_output1 = self.y_train[self.y_train == 1]
        self.y_train_output0 = self.y_train[self.y_train == 0]
        self.y_test_output1 = self.y_test[self.y_test == 1]
        self.y_test_output0 = self.y_test[self.y_test == 0]
        print("y_train_output1 type: {} and shape: {}".format(type(self.y_train_output1), self.y_train_output1.shape))
        print("y_test_output1 type: {} and shape: {}".format(type(self.y_test_output1), self.y_test_output1.shape))
        print("y_train_output0 type: {} and shape: {}".format(type(self.y_train_output0), self.y_train_output0.shape))
        print("y_test_output0 type: {} and shape: {} \n".format(type(self.y_test_output0), self.y_test_output0.shape))

        plt.subplots(figsize=(self.fig_width, self.fig_height))
        plt.bar([1, 2], [self.y_train_output0.shape[0], self.y_test_output0.shape[0]],
                color='r', width=self.bar_width, edgecolor='black', label='class=0')
        plt.bar([1 + self.bar_width, 2 + self.bar_width], [self.y_train_output1.shape[0], self.y_test_output1.shape[0]],
                color='b', width=self.bar_width, edgecolor='black', label='class=1')
        plt.xticks([1 + self.bar_width / 2, 2 + self.bar_width / 2],
                   ['Train data', 'Test data'], ha='center')
        plt.text(1 - self.bar_width / 4, self.y_train_output0.shape[0] + 100,
                 str(self.y_train_output0.shape[0]), fontsize=20)
        plt.text(1 + 3 * self.bar_width / 4, self.y_train_output1.shape[0] + 100,
                 str(self.y_train_output1.shape[0]), fontsize=20)
        plt.text(2 - self.bar_width / 4, self.y_test_output0.shape[0] + 100,
                 str(self.y_test_output0.shape[0]), fontsize=20)
        plt.text(2 + 3 * self.bar_width / 4, self.y_test_output1.shape[0] + 100,
                 str(self.y_test_output1.shape[0]), fontsize=20)
        plt.title('Output class distribution between train and test datasets', fontsize=24)
        plt.xlabel('Concepts', fontweight='bold', fontsize=14)
        plt.ylabel('Count train/test class cases', fontweight='bold', fontsize=14)
        plt.legend()
        plt.grid()
        plt.savefig('Count_class_cases.png', bbox_inches='tight')
        plt.clf()

    def data_scaling(self, algorithm):
        """Scaling data to normalization or standardization"""
        if algorithm.lower() == 'norm':
            scaler = MinMaxScaler()
        elif algorithm.lower() == 'standard':
            scaler = StandardScaler()
        else:
            print('Algorithm not correct')
            return None
        scaler.fit(self.X_train)
        self.X_train_scaled = scaler.transform(self.X_train)
        self.X_test_scaled = scaler.transform(self.X_test)
        print("X_train_scaled type: {} and shape: {}".format(type(self.X_train_scaled), self.X_train_scaled.shape))
        print("X_test_scaled type: {} and shape: {} \n".format(type(self.X_test_scaled), self.X_test_scaled.shape))

    def apply_algorithm(self, algorithm, params):
        """Apply the machine learning algorithm to the train and test datasets"""
        time0 = time.time()
        if algorithm.lower() == 'knn':
            model = KNeighborsClassifier()
        elif algorithm.lower() == 'logreg':
            model = LogisticRegression(random_state=0)
        elif algorithm.lower() == 'linearsvc':
            model = LinearSVC(random_state=0)
        elif algorithm.lower() == 'naivebayes':
            model = GaussianNB()
        elif algorithm.lower() == 'tree':
            model = DecisionTreeClassifier(random_state=0)
        elif algorithm.lower() == 'forest':
            model = RandomForestClassifier(random_state=0)
        elif algorithm.lower() == 'gradient':
            model = GradientBoostingClassifier(random_state=0)
        elif algorithm.lower() == 'svm':
            model = SVC(random_state=0)
        elif algorithm.lower() == 'mlp':
            model = MLPClassifier(random_state=0)
        else:
            return None
        for key, value in params.items():
            setattr(model, key, value)
        print('SCORE WITH {} ALGORITHM AND PARAMS {}\n'.format(algorithm, params))
        model.fit(self.X_train, self.y_train)
        time1 = time.time()
        unscaled_model_time = round(time1 - time0, 4)
        unscaled_train_score = round(model.score(self.X_train, self.y_train), 4)
        unscaled_test_score = round(model.score(self.X_test, self.y_test), 4)
        print('Unscaled modeling time [seconds]: {}'.format(unscaled_model_time, 4))
        print('Unscaled TRAIN dataset: {}'.format(unscaled_train_score, 4))
        print('Unscaled TEST dataset: {}'.format(unscaled_test_score, 4))
        time2 = time.time()
        unscaled_predict_time = round(time2 - time1, 4)
        print('Unscaled predicting time [seconds]: {}\n'.format(unscaled_predict_time, 4))
        model.fit(self.X_train_scaled, self.y_train)
        time3 = time.time()
        scaled_model_time = round(time3 - time2, 4)
        scaled_train_score = round(model.score(self.X_train_scaled, self.y_train), 4)
        scaled_test_score = round(model.score(self.X_test_scaled, self.y_test), 4)
        print('Scaled modeling time [seconds]: {}'.format(scaled_model_time, 4))
        print('Scaling TRAIN dataset: {}'.format(scaled_train_score, 4))
        print('Scaling TEST dataset: {}'.format(scaled_test_score, 4))
        time4 = time.time()
        scaled_predict_time = round(time4 - time3, 4)
        print('Scaled predicting time [seconds]: {}\n\n'.format(scaled_predict_time, 4))
        results = np.array([[unscaled_model_time, unscaled_predict_time, unscaled_train_score, unscaled_test_score,
                             scaled_model_time, scaled_predict_time, scaled_train_score, scaled_test_score]])
        self.out_data = np.append(self.out_data, results, axis=0)
        self.method.append(algorithm)
        self.parameters.append(params)

    def write_results_excel_file(self, name):
        """Write the simulation results in an output excel file"""
        # Create excel file with the corresponding sheets
        sheets = []
        wb = Workbook()
        sheets.append(wb.active)
        sheets[0].title = 'SIMULATION RESULTS'
        # Define column width
        for column in range(1, 11):
            column_char = str(chr(64 + column))
            if column == 2:
                sheets[0].column_dimensions[column_char].width = 60
            else:
                sheets[0].column_dimensions[column_char].width = 20
        # Write headers
        header = ['Algorithm', 'Params', 'Unscaled Model Time', 'Unscaled Predict Time', 'Unscaled Train Score',
                  'Unscaled Test Score', 'Scaled Model Time', 'Scaled Predict Time', 'Scaled Train Score',
                  'Scaled Test Score']
        for i in range(len(header)):
            sheets[0].cell(1, i + 1).value = header[i]
        # Write algorithms
        for i in range(len(self.method)):
            sheets[0].cell(i + 2, 1).value = self.method[i]
        # Write parameters
        for i in range(len(self.parameters)):
            str_params = ''
            for key, value in self.parameters[i].items():
                str_params += ' ' + key + '=' + str(value)
            sheets[0].cell(i + 2, 2).value = str_params
        # Write data in excel sheet
        for i in range(self.out_data.shape[0]):
            for j in range(self.out_data.shape[1]):
                sheets[0].cell(i + 2, j + 3).value = self.out_data[i, j]
        try:
            wb.save(name)
        except PermissionError:
            sys.exit('ERROR: Excel file open. Please close it to be modified')
